# -*- coding: utf8-*-
from __future__ import unicode_literals
from difflib import SequenceMatcher
import youtube_dl
import os
import openpyxl, time

def youtube_crawling_():
    exel_file = openpyxl.load_workbook('resource/영역_title_keyword_id_190819_original.xlsx', data_only=True)
    excel_sheet = exel_file.get_sheet_by_name("Sheet1")

    # ydl_opts = {
    #         'format': 'bestaudio/best',
    #         'extractaudio': False,
    #         # 'audioformat': "mp3",
    #         'postprocessors': [{
    #         'key': 'FFmpegExtractAudio',
    #         'preferredcodec': 'wav',
    #         'preferredquality': '360'}],
    #         # 'noplaylist': True,
    #         # 'listformats': True
    #         'outtmpl': '%(id)s)%(title)s.%(ext)s',
    #         # 'outtmpl': '%(id)s.%(ext)s',
    #         'writesubtitles': True,
    #         }

    ydl_opts = {
        'extractaudio': False,
        'skip_download': True,
        'outtmpl': '%(id)s)%(title)s.%(ext)s',
        'subtitleslangs': ["ko"],
        'writesubtitles': True,
        'writeautomaticsub': True
    }

    # 다운 받을 영상이 저장될 디렉터리
    # os.chdir('./youtube_audio/korean')
    os.chdir('/media/hyunseokahn/대용량/audiofile')

    exist_id = []
    for r_seq_1, row_1 in enumerate(excel_sheet.rows):
        if r_seq_1 > 0:
            if row_1[0].value != "회화" and row_1[0].value != "영어":
                youtube_id = row_1[3].value
                # 현재 폴더에 있는 영상들의 ID 값들을 불러와서 비교하고, 그 이후부터 진행
                if youtube_id in exist_id:
                    continue
                try:
                    with youtube_dl.YoutubeDL(ydl_opts) as ydl:
                        # result = ydl.extract_info('https://www.youtube.com/watch?v=QXvSrH4Tcek', download=False)
                        ydl.download(['https://www.youtube.com/watch?v=' + youtube_id])
                except Exception as ex:
                    print(str(youtube_id) + " - 오디오 파일을 다운 받을 수 없습니다. : " + str(ex))


def load_subscribe():
    cress_file = openpyxl.load_workbook('resource/영역_title_keyword_id_190819_notnan_space.xlsx', data_only=True)
    cress_sheet = cress_file.get_sheet_by_name("Sheet1")

    path_dir = '/media/hyunseokahn/대용량/audiofile/'
    file_list = os.listdir(path_dir)
    file_list_vtt = [file for file in file_list if file.endswith(".vtt")]
    print(len(file_list_vtt))

    list_a = []
    flag = False
    tmp_text = ''
    for r_seq_1, row_1 in enumerate(cress_sheet.rows):
        if r_seq_1 > 0:
            list_a.append(row_1[3].value)
            for name in file_list_vtt:
                if row_1[3].value == name.split(")")[0]:
                    file_text = ''
                    with open('/media/hyunseokahn/대용량/audiofile/' + name, 'r', encoding='utf-8') as file:
                        lines = file.readlines()
                        for line in lines:
                            # print(line)
                            if flag == True:
                                if not "[음악]" in line and not "[박수]" in line and not "[웃음]" in line and not "[박수 갈채]" in line \
                                        and not line.strip() == "":
                                    if not tmp_text == line:
                                        tmp_text = line
                                        file_text += line.replace("\n", " ")
                                        # print(line)
                                flag = False
                                continue
                            if line[:3] == "00:":
                                flag = True
                    print(file_text)
                    print("---------------------------------------------------------------------------")
                    cress_sheet.cell(row=r_seq_1 + 1, column=5, value=file_text)

    cress_file.save('test.xlsx')

    # w_count = {}
    # for lst in list_a:
    #     try:
    #         w_count[lst] += 1
    #     except:
    #         w_count[lst] = 1


def evaluation_stt_result():
    filenames = os.listdir("result/유튜브 결과 파일/korean/")
    flag = 0
    google_mean_ratio = 0
    clover_mean_ratio = 0
    ibm_mean_ratio = 0
    microsoft_mean_ratio = 0
    google_count = 0
    clover_count = 0
    ibm_count = 0
    microsoft_count = 0

    for filename in filenames:
        if "google" in filename:
            file = open("result/유튜브 결과 파일/korean/" + filename, 'r')
            org_text = file.read()
            flag = 1

        elif "clover" in filename:
            file = open("result/유튜브 결과 파일/korean/" + filename, 'r', encoding='utf8')
            org_text = file.read()
            flag = 2

        elif "ibm" in filename:
            file = open("result/유튜브 결과 파일/korean/" + filename, 'r')
            org_text = file.read()
            flag = 3

        elif "microsoft" in filename:
            file = open("result/유튜브 결과 파일/korean/" + filename, 'r')
            org_text = file.read()
            flag = 4

        if flag == 1:
            newfile = filename.replace("google_", "")
            # print(filename)
            testfile = open("result/유튜브 결과 파일/korean/" + newfile, 'r', encoding='utf8')
            test_text = testfile.read()
            ratio = SequenceMatcher(None, org_text, test_text).ratio()
            google_mean_ratio += ratio
            google_count += 1

        elif flag == 2:
            newfile = filename.replace("clover_", "")
            # print(filename)
            testfile = open("result/유튜브 결과 파일/korean/" + newfile, 'r', encoding='utf8')
            test_text = testfile.read()
            ratio = SequenceMatcher(None, org_text, test_text).ratio()
            clover_mean_ratio += ratio
            clover_count += 1

        elif flag == 3:
            newfile = filename.replace("ibm_", "")
            # print(filename)
            testfile = open("result/유튜브 결과 파일/korean/" + newfile, 'r', encoding='utf8')
            test_text = testfile.read()
            ratio = SequenceMatcher(None, org_text, test_text).ratio()
            ibm_mean_ratio += ratio
            ibm_count += 1

        elif flag == 4:
            newfile = filename.replace("microsoft_", "")
            # print(filename)
            testfile = open("result/유튜브 결과 파일/korean/" + newfile, 'r', encoding='utf8')
            test_text = testfile.read()
            ratio = SequenceMatcher(None, org_text, test_text).ratio()
            microsoft_mean_ratio += ratio
            microsoft_count += 1

        # print(ratio)

    print(" : " + str(clover_mean_ratio / clover_count))
    print(" : " + str(google_mean_ratio / google_count))
    print(" : " + str(ibm_mean_ratio / ibm_count))
    print(" : " + str(microsoft_mean_ratio / microsoft_count))


try:
    import azure.cognitiveservices.speech as speechsdk
except ImportError:
    print("""
    Importing the Speech SDK for Python failed.
    Refer to
    https://docs.microsoft.com/azure/cognitive-services/speech-service/quickstart-python for
    installation instructions.
    """)
    import sys

    sys.exit(1)


def speech_recognize_continuous_from_file():
    """performs continuous speech recognition with input from an audio file"""
    # <SpeechContinuousRecognitionWithFile>

    weatherfilename = "download/workman2.wav"

    speech_config = speechsdk.SpeechConfig(subscription=speech_key, region=service_region, speech_recognition_language="ko-KR")
    audio_config = speechsdk.audio.AudioConfig(filename=weatherfilename)

    speech_recognizer = speechsdk.SpeechRecognizer(speech_config=speech_config, audio_config=audio_config)

    done = False

    def stop_cb(evt):
        """callback that stops continuous recognition upon receiving an event `evt`"""
        print('CLOSING on {}'.format(evt))
        speech_recognizer.stop_continuous_recognition()
        nonlocal done
        done = True

    # Connect callbacks to the events fired by the speech recognizer
    speech_recognizer.recognizing.connect(lambda evt: print('RECOGNIZING: {}'.format(evt)))

    # Start continuous speech recognition
    speech_recognizer.start_continuous_recognition()
    while not done:
        time.sleep(.5)
    # </SpeechContinuousRecognitionWithFile>


from pydub import AudioSegment

def cutting_audio():
    file_name = "Ten in the Bed _ Nursery Rhymes and Baby Songs"
    # Open file
    # song = AudioSegment.from_wav('youtube_audio/eng/' + file_name + '.wav')
    song = AudioSegment.from_wav('youtube_audio/korean/' + file_name + '.wav')

    print("재생 시간 : " + str(int(len(song) / 1000)) + " 초")

    # Slice audios
    # pydub는 milliseconds 단위를 사용
    ten_seconds = 10 * 1000
    one_min = ten_seconds * 6

    start_time = ten_seconds * 5
    end_time = ten_seconds * 8

    first_10_seconds = song[start_time:end_time]
    # last_5_seconds = song[-5000:]

    # up/down volumn
    beginning = first_10_seconds + 6

    # Save the result
    # can give parameters-quality, channel, etc
    beginning = beginning.set_channels(1)
    beginning.export(
        'youtube_audio/eng/' + file_name + '_' + str(int(start_time / 1000)) + '_' + str(int(end_time / 1000)) + '.wav',
        format='wav')
    # beginning.export('youtube_audio/korean/' + file_name + '_' + str(int(start_time/1000)) + '_' + str(int(end_time/1000)) + '.wav', format='wav')

    # beginning.exoprt('result.flac', format='flac', parameters=["-q:a", "10", "-ac", "1"])

    # sound = AudioSegment.from_wav("/path/to/file.wav")
    # sound = sound.set_channels(1)
    # sound.export("/output/path.wav", format="wav")


import pyaudio
import wave

def streaming_recoder():
    FORMAT = pyaudio.paInt16
    CHANNELS = 1  # only mono
    RATE = 16000
    CHUNK = 1024  # 확인 필요
    RECORD_SECONDS = 10  # 10초 녹음

    WAVE_OUTPUT_FILENAME = "file.flac"

    audio = pyaudio.PyAudio()

    # start Recording
    stream = audio.open(format=FORMAT, channels=CHANNELS,
                        rate=RATE, input=True,
                        frames_per_buffer=CHUNK)
    print("recording...")
    frames = []

    for i in range(0, int(RATE / CHUNK * RECORD_SECONDS)):
        data = stream.read(CHUNK)
        frames.append(data)
    print("finished recording")

    # stop Recording
    stream.stop_stream()
    stream.close()
    audio.terminate()

    waveFile = wave.open(WAVE_OUTPUT_FILENAME, 'wb')
    waveFile.setnchannels(CHANNELS)
    waveFile.setsampwidth(audio.get_sample_size(FORMAT))
    waveFile.setframerate(RATE)
    waveFile.writeframes(b''.join(frames))
    waveFile.close()